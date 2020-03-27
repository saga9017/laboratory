import torch
from sklearn import datasets
import random

n_pts = 6000
X, y = datasets.make_circles(n_samples=n_pts, random_state=123, noise=0.25, factor=0.3)
x_data = torch.tensor(X).float()
y_data = torch.tensor(y.reshape(6000, 1)).float()

x_new_data=[]

for indx, i in enumerate(x_data):
    x_new_data.append([indx]+i.tolist())

random.shuffle(x_new_data)

x_train=[]
y_train=[]
for i in x_new_data[:5000]:
    x_train.append(i[1:])
    y_train.append(y_data[i[0]])

x_test=[]
y_test=[]
for i in x_new_data[5000:5500]:
    x_test.append(i[1:])
    y_test.append(y_data[i[0]])

x_val=[]
y_val=[]
for i in x_new_data[5500:]:
    x_val.append(i[1:])
    y_val.append(y_data[i[0]])

print(len(x_train))
print(len(y_train))
print(len(x_test))
print(len(x_test))
print(len(x_val))
print(len(x_val))
########################################################################################################################
from openpyxl import Workbook

write_wb = Workbook()

# 이름이 있는 시트를 생성
# write_ws = write_wb.create_sheet('생성시트')

# Sheet1에다 입력
write_ws = write_wb.active
write_ws['A1'] = 'X1'
write_ws['B1'] = 'X2'
write_ws['C1'] = 'Class'

# 행 단위로 추가
for x, y in zip(x_train, y_train):
    write_ws.append([x[0], x[1], y.item()])

write_wb.save('train_set.xlsx')

########################################################################################################################

write_wb2 = Workbook()

# 이름이 있는 시트를 생성
# write_ws = write_wb.create_sheet('생성시트')

# Sheet1에다 입력
write_ws2 = write_wb2.active
write_ws2['A1'] = 'X1'
write_ws2['B1'] = 'X2'
write_ws2['C1'] = 'Class'

# 행 단위로 추가
for x, y in zip(x_test, y_test):
    write_ws2.append([x[0], x[1], y.item()])

write_wb2.save('test_set.xlsx')

########################################################################################################################

write_wb3 = Workbook()

# 이름이 있는 시트를 생성
# write_ws = write_wb.create_sheet('생성시트')

# Sheet1에다 입력
write_ws3 = write_wb3.active
write_ws3['A1'] = 'X1'
write_ws3['B1'] = 'X2'
write_ws3['C1'] = 'Class'

# 행 단위로 추가
for x, y in zip(x_val, y_val):
    write_ws3.append([x[0], x[1], y.item()])

write_wb3.save('val_set.xlsx')
